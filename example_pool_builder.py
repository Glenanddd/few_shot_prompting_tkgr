"""
Offline builder for a dynamic example pool (JSONL).

This script constructs a local "example pool" from:
1) sampled_path/{dataset}/confidence_concrete.json
   - random-walk closed paths with confidence (acts as example *inputs*)
2) gen_rules_iteration/{dataset}/evaluation/train/0_confidence_concrete.json
   - 0-th iteration generated rules after clean/eval (acts as example *outputs*)

Each JSONL line is an example item:
{
  "ex_id": "<head_rel_name>#<idx>",
  "head_rel_name": "...",
  "paths": [str, ...],
  "rules": [str, ...],
  "path_rels": [str, ...]
}

NOTE: This is *offline preprocessing* only. Do NOT run this during Iteration_reasoning.py.

In addition to the JSONL pool, this script also writes a companion statistics workbook:
  <out_basename>_stats.xlsx
with multiple sheets (Summary/PerHead/PerItem/Dropped/TokenFreq/Glossary) to help you audit
coverage and quality. This requires `openpyxl`.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import statistics
from collections import Counter, defaultdict
from typing import Dict, Iterable, List, Tuple
from utils_windows_long_path import maybe_windows_long_path

# Extract a predicate name from temporal atoms like:
#   Return,_release_person(s)(X0,X1,T0)
# The relation name itself may contain parentheses "(s)", so we anchor on the *argument* pattern.
_REL_TOKEN_RE = re.compile(r"([\w\s'\-.,\(\)/]+)\(\w+,\s*\w+,\s*\w+\)")

# Strip leading "conf / supp / ..." fields from "abstract_rule" strings:
#   "0.750000     6     8  Make_statement(X0,X1,T3)<-..."
#
# IMPORTANT:
# - We must avoid capturing the numeric prefix (conf/supp/body_supp), so the rule core must start from a
#   relation name that begins with a non-digit (typical in ICEWS: letters / inv_*).
_RULE_CORE_RE = re.compile(r"([A-Za-z_][\w\s'\-.,\(\)/]*\(\w+,\s*\w+,\s*\w+\)<-.*)$")

# Fallback: directly strip "<conf> <rule_supp> <body_supp>" prefix if the core regex fails.
_METRIC_PREFIX_RE = re.compile(
    r"^\s*[-+]?(?:\d+(?:\.\d+)?|\.\d+)(?:[eE][-+]?\d+)?\s+\d+\s+\d+\s+"
)


def _normalize_ws(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def strip_rule_metrics(abstract_rule: str) -> str:
    """
    Remove the leading numeric fields (conf/supp/etc.) and keep only the rule string.
    """
    abstract_rule = abstract_rule.strip()
    match = _RULE_CORE_RE.search(abstract_rule)
    if match:
        return match.group(1).strip()
    # Fallback: best-effort strip the numeric prefix, if present.
    stripped = _METRIC_PREFIX_RE.sub("", abstract_rule).strip()
    return stripped or abstract_rule


def extract_relation_tokens(rule_text: str) -> List[str]:
    """
    Similarity feature (interpretable, local): relation-name tokens extracted from rule *bodies*.
    We use only the body-part (after '<-') to avoid trivial matching by the head relation name.
    """
    body = rule_text
    if "<-" in rule_text:
        body = rule_text.split("<-", 1)[1]
    tokens = [_normalize_ws(t) for t in _REL_TOKEN_RE.findall(body)]
    return [t for t in tokens if t]


def _collect_rules_by_head(json_path: str) -> Dict[str, List[Tuple[float, str]]]:
    """
    Load a dict-of-lists JSON and return: head_rel_name -> [(conf, rule_str), ...]
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    by_head: Dict[str, List[Tuple[float, str]]] = defaultdict(list)
    for _head_key, items in data.items():
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            head_name = item.get("head_rel_name")
            if not head_name:
                continue
            conf = float(item.get("conf", 0.0) or 0.0)
            abstract_rule = item.get("abstract_rule") or ""
            rule_core = strip_rule_metrics(str(abstract_rule))
            if not rule_core:
                continue
            by_head[head_name].append((conf, rule_core))
    return by_head


def _chunk_list(xs: List, chunk_size: int, num_chunks: int) -> List[List]:
    chunks: List[List] = []
    for i in range(num_chunks):
        chunk = xs[i * chunk_size : (i + 1) * chunk_size]
        chunks.append(chunk)
    return chunks


def _count_body_atoms(rule_text: str) -> int:
    """
    Count the number of body atoms in a rule string like:
      Head(...)<-R1(...)&R2(...)&...
    """
    if "<-" not in rule_text:
        return 0
    body = rule_text.split("<-", 1)[1].strip()
    if not body:
        return 0
    return len([x for x in body.split("&") if x.strip()])


def _safe_mean(xs: List[float]):
    return (sum(xs) / len(xs)) if xs else None


def _safe_median(xs: List[float]):
    return statistics.median(xs) if xs else None


def _safe_min(xs: List[float]):
    return min(xs) if xs else None


def _safe_max(xs: List[float]):
    return max(xs) if xs else None


def _to_table(rows: List[dict], columns: List[str]) -> List[List[object]]:
    table: List[List[object]] = [columns]
    for r in rows:
        table.append([r.get(c) for c in columns])
    return table


def _sanitize_sheet_name(name: str) -> str:
    # Excel constraints: <= 31 chars, cannot contain : \ / ? * [ ]
    name = re.sub(r"[:\\\\/?*\\[\\]]", "_", name)
    return name[:31] if len(name) > 31 else name


def write_stats_xlsx(stats_tables: List[Tuple[str, List[List[object]]]], out_xlsx_path: str) -> None:
    """
    Write multiple 2D tables into an .xlsx file via openpyxl.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise ImportError(
            "openpyxl is required to write stats XLSX. Please install it (e.g., `pip install openpyxl`)."
        ) from e

    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx_path)) or ".", exist_ok=True)

    wb = Workbook()
    # Remove the default worksheet created by openpyxl.
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_alignment = Alignment(vertical="top", wrap_text=False)
    normal_alignment = Alignment(vertical="top", wrap_text=False)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for name, table in stats_tables:
        ws = wb.create_sheet(title=_sanitize_sheet_name(name))
        for row in table:
            ws.append(row)

        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_alignment

        # Wrap long text only in the Glossary sheet.
        body_alignment = wrap_alignment if ws.title == "Glossary" else normal_alignment
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = body_alignment

        # Column widths (scan only the first N rows for speed).
        max_scan_rows = min(ws.max_row, 300)
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for r in range(1, max_scan_rows + 1):
                v = ws.cell(row=r, column=col).value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[get_column_letter(col)].width = min(max(8, max_len + 2), 80)

    wb.save(out_xlsx_path)


def build_glossary_table(
    *,
    summary_metrics: List[str],
    per_head_columns: List[str],
    per_item_columns: List[str],
    dropped_columns: List[str],
    token_columns: List[str],
    E: int,
    P: int,
    R: int,
) -> List[List[object]]:
    """
    Build a "Glossary" sheet:
    - Explain every metric/field in the workbook
    - Provide a small example to make it easy for others to understand
    """

    def example_int(n: int) -> str:
        return f"例：{n}"

    def example_float(x: float) -> str:
        return f"例：{x}"

    def describe(sheet: str, field: str) -> Tuple[str, str, str, str]:
        # Returns: (meaning, how_computed, example, how_to_read)
        if sheet == "Summary":
            if field == "dataset":
                return (
                    "数据集名称。",
                    "命令行参数 --dataset。",
                    "例：icews14",
                    "用于定位输入/输出路径以及复现实验设置。",
                )
            if field in {"sampled_paths_json", "generated_rules_json", "out_jsonl", "stats_xlsx"}:
                return (
                    "文件路径。",
                    "构建时使用/写入的实际路径。",
                    "例：prompt/icews14/example_pool.jsonl",
                    "用于追踪本次构建使用的数据来源与产物位置。",
                )
            if field == "strict_full_chunk_filter":
                return (
                    "是否启用“严格满配过滤”。",
                    "若为 True，则 p_chunk 必须满足 len==P 且 r_chunk 必须满足 len==R，否则该 item 直接丢弃。",
                    "例：P=8,R=6，某 chunk 只有 7 条 path => 丢弃",
                    "该开关越严格，池子越“规整”，但覆盖率可能下降。",
                )
            if field == "token_extraction_body_only":
                return (
                    "检索 token 是否只来自 rule body（'<-' 右侧）。",
                    "True 表示 token 抽取仅在 body 上做，避免因 head 名称导致“虚假相似”。",
                    "例：只抽取 A(...)&B(...) 中的 A,B，不抽 head",
                    "这会让检索更关注路径模式而非 head 名字本身。",
                )
            if field == "top_n_paths":
                return (
                    "每个 head 最多使用多少条 sampled paths（按 conf 降序截断）。",
                    "n_paths_used = min(top_n_paths, n_paths_total)。",
                    f"例：top_n_paths=40，n_paths_total=120 => n_paths_used=40",
                    "该值过小会导致很多 head 无法凑够 E*P；过大则会增大构建时间但通常无害。",
                )
            if field == "top_n_rules":
                return (
                    "每个 head 最多使用多少条生成规则（按 conf 降序截断）。",
                    "n_rules_used = min(top_n_rules, n_rules_total)。",
                    f"例：top_n_rules=30，n_rules_total=12 => n_rules_used=12",
                    "该值过小会导致很多 head 无法凑够 E*R。",
                )
            if field == "num_items_per_head(E)":
                return (
                    "每个 head 目标最多构建的 item 数（E）。",
                    "paths/rules 各自切成 E 个 chunk，然后按 chunk_idx 配对。",
                    f"例：E={E}，每个 head 最多 {E} 个 item",
                    "E 越大，池子更大、覆盖更多模式，但也可能增加检索噪声与 prompt 长度。",
                )
            if field == "paths_per_item(P)":
                return (
                    "每个 item 中包含的 sampled paths 数（P）。",
                    "paths chunk 大小固定为 P。",
                    f"例：P={P}，item.paths 长度应为 {P}",
                    "P 越大，每个示例输入更丰富，但也更占 prompt token。",
                )
            if field == "rules_per_item(R)":
                return (
                    "每个 item 中包含的 generated rules 数（R）。",
                    "rules chunk 大小固定为 R。",
                    f"例：R={R}，item.rules 长度应为 {R}",
                    "R 越大，每个示例输出更丰富，但也更占 prompt token。",
                )
            if field == "required_paths(E*P)":
                return (
                    "构建满配 E 个 item 所需的 paths 数。",
                    "required_paths = E * P。",
                    f"例：E={E}, P={P} => required_paths={E * P}",
                    "若 n_paths_used < required_paths，则必然无法产满 E 个 item。",
                )
            if field == "required_rules(E*R)":
                return (
                    "构建满配 E 个 item 所需的 rules 数。",
                    "required_rules = E * R。",
                    f"例：E={E}, R={R} => required_rules={E * R}",
                    "若 n_rules_used < required_rules，则必然无法产满 E 个 item。",
                )
            if field == "heads_in_paths":
                return (
                    "sampled_paths_json 中出现的 head 数。",
                    "对 JSON 的 head_rel_name 取集合大小。",
                    example_int(100),
                    "用于判断 sampled_paths 侧覆盖面。",
                )
            if field == "heads_in_rules":
                return (
                    "generated_rules_json 中出现的 head 数。",
                    "对 JSON 的 head_rel_name 取集合大小。",
                    example_int(100),
                    "用于判断生成规则侧覆盖面。",
                )
            if field == "heads_intersection":
                return (
                    "两侧都有数据的 head 数（交集）。",
                    "heads_intersection = |heads_in_paths ∩ heads_in_rules|。",
                    example_int(80),
                    "只有交集里的 head 才可能构建出 item。",
                )
            if field == "candidate_items_total(heads*E)":
                return (
                    "理论最大 item 数（不考虑缺失/过滤）：交集 head 数 * E。",
                    "candidate_items_total = heads_intersection * E。",
                    f"例：heads_intersection=80,E={E} => {80 * E}",
                    "用于对比 items_written 看“损失率”。",
                )
            if field == "items_written":
                return (
                    "最终写入 JSONL 的 item 数。",
                    "对所有 head 的 full chunks 计数（且通过严格过滤）。",
                    example_int(1754),
                    "越大代表池子更大；但也要结合覆盖率与 token 多样性看质量。",
                )
            if field == "items_dropped_empty":
                return (
                    "因为 chunk 为空而丢弃的 item 数（paths 或 rules 其中一侧为 0 条）。",
                    "统计 dropped_rows 中 reason=empty_* 的数量。",
                    example_int(10),
                    "通常表示某些 head 的 top_n 之后严重不足，或交集 head 里一侧数据很稀疏。",
                )
            if field == "items_dropped_partial":
                return (
                    "因为 chunk 不满配而丢弃的 item 数（len<P 或 len<R）。",
                    "统计 dropped_rows 中 reason=partial_* 的数量。",
                    example_int(100),
                    "该值越大说明 P/R/E 或 top_n_* 设得过严，覆盖率会下降。",
                )
            if field == "heads_with>=1_item":
                return (
                    "至少产出 1 个 full item 的 head 数。",
                    "对交集 head 统计 items_full_written>=1。",
                    example_int(200),
                    "越大说明池子覆盖越广。",
                )
            if field == "heads_with_full_E_items":
                return (
                    "产满 E 个 full item 的 head 数。",
                    "对交集 head 统计 items_full_written>=E。",
                    example_int(50),
                    "越大说明当前 E/P/R/top_n_* 更容易满足满配。",
                )
            if field in {"heads_paths_used>=E*P", "heads_rules_used>=E*R", "heads_both_used>=E*P_and_E*R"}:
                return (
                    "在 top_n 截断后仍满足满配所需数量的 head 数（供给侧是否足够）。",
                    "分别统计 n_paths_used>=E*P、n_rules_used>=E*R、两者同时满足。",
                    example_int(120),
                    "这是调参的关键：若该值很低，说明 P/R/E/top_n_* 组合对数据不友好。",
                )
            if field in {"items_per_head_mean", "items_per_head_median"}:
                return (
                    "每个 head 的 full item 数的均值/中位数。",
                    "统计交集 head 的 items_full_written。",
                    example_float(2.3),
                    "若均值远小于 E，说明很多 head 无法满配。",
                )
            if field == "unique_tokens_in_kept_items":
                return (
                    "保留下来的 items 中，抽取到的 relation token 的去重数量。",
                    "对所有写入 item 的 path_tokens 取集合大小。",
                    example_int(500),
                    "越大说明检索特征更丰富，通常有利于“相似+多样”。",
                )
            if field.startswith("heads_with_items="):
                return (
                    "产出恰好 i 个 full item 的 head 数（i 由字段名给出）。",
                    "从 PerHead 的 items_full_written 统计直方图。",
                    f"例：heads_with_items=3 表示有若干 head 产出 3 个 item",
                    "用于快速观察覆盖分布是否极度不均衡。",
                )

        if sheet == "PerHead":
            if field == "head_rel_name":
                return (
                    "规则头关系名称（rule head relation）。",
                    "来自输入 JSON 的 head_rel_name。",
                    "例：Make_statement",
                    "用于定位哪个 head 的样例池质量更好/更差。",
                )
            if field in {"in_paths", "in_rules"}:
                return (
                    "该 head 是否出现在对应输入文件中。",
                    "head_rel_name 是否存在于对应 JSON 的 key 集合中。",
                    "例：in_paths=True, in_rules=False",
                    "若任一为 False，则该 head 不会产生 item。",
                )
            if field in {"n_paths_total", "n_rules_total"}:
                return (
                    "该 head 在原始输入 JSON 中的规则条数。",
                    "对应 head_rel_name 的列表长度。",
                    example_int(120),
                    "用于判断数据量是否足够支撑满配 E 个 item。",
                )
            if field in {"n_paths_used", "n_rules_used"}:
                return (
                    "该 head 实际用于构建的条数（top_n 截断后）。",
                    "min(top_n_*, n_*_total)。",
                    example_int(40),
                    "如果 used 远小于 required(E*P/E*R)，则无法满配。",
                )
            if field in {"required_paths(E*P)", "required_rules(E*R)"}:
                return (
                    "该 head 满配 E 个 item 的需求量（与 Summary 相同）。",
                    "required_paths=E*P，required_rules=E*R。",
                    f"例：E={E},P={P},R={R} => {E*P}/{E*R}",
                    "用于快速对比 n_*_used 是否足够。",
                )
            if field in {"full_path_chunks", "full_rule_chunks"}:
                return (
                    "在 strict 满配下，该 head 在某一侧最多能形成多少个“满配 chunk”。",
                    "full_path_chunks = n_paths_used // P；full_rule_chunks = n_rules_used // R。",
                    f"例：n_paths_used=33,P={P} => {33 // P}",
                    "最终可写入的 item 数受两侧较小值限制。",
                )
            if field in {"items_non_empty", "items_full_written", "items_target_E"}:
                return (
                    "该 head 的 chunk 配对情况：非空配对数 / 满配写入数 / 目标 E。",
                    "items_full_written = 通过 strict 过滤的 chunk 对数量；items_target_E=E。",
                    f"例：E={E}，items_full_written=2 表示该 head 只有 2 个 item",
                    "items_full_written 越小，说明该 head 对池子贡献较少。",
                )
            if field in {"dropped_empty", "dropped_partial"}:
                return (
                    "该 head 内部被丢弃的 chunk 对数量（按原因分）。",
                    "统计 Dropped sheet 中同 head 的记录数量（按 reason 聚合）。",
                    example_int(3),
                    "用于定位是“缺数据导致空”还是“差一点点不满配导致丢弃”。",
                )
            if field == "bottleneck":
                return (
                    "该 head 无法满配的主要瓶颈类型。",
                    "基于 full_path_chunks/full_rule_chunks 以及缺失侧给出：missing_* / paths_short / rules_short / both_short / ok_full。",
                    "例：paths_short",
                    "用来决定应该调大 top_n_paths 还是 top_n_rules，或调小 P/R/E。",
                )
            if field.startswith("paths_conf_") or field.startswith("rules_conf_"):
                return (
                    "该 head 的 conf 统计（基于 top_n 截断后的 used 列表）。",
                    "对 used 的 conf 列做 mean/median/min/max。",
                    "例：conf_mean=0.42",
                    "可用于判断：留下的数据是否整体置信度偏低（可能影响示例质量）。",
                )
            if field.endswith("_body_atoms_mean_used"):
                return (
                    "该 head 的平均 body 原子数量（规则体中 & 分隔的关系数）。",
                    "count_body_atoms(rule)=len(body.split('&'))；对 used 取均值。",
                    "例：A& B& C => body_atoms=3",
                    "值越大表示路径/规则更长、更复杂，可能更占 prompt token。",
                )
            if field == "paths_unique_tokens_used":
                return (
                    "该 head 的 used paths 中抽取到的关系 token 去重数。",
                    "对 used paths 的 body 抽取 token 后取去重。",
                    example_int(25),
                    "越大说明该 head 的路径关系更丰富，检索特征更强。",
                )
            if field == "paths_empty_token_ratio_used":
                return (
                    "token 抽取失败比例：used paths 中有多少条抽不到 token。",
                    "empty_ratio = empty_token_paths / n_paths_used。",
                    "例：2/40=0.05",
                    "若该值很高，说明规则格式可能异常或 regex 不匹配，检索会变弱。",
                )

        if sheet == "PerItem":
            if field == "ex_id":
                return (
                    "样例 item 唯一 ID。",
                    "ex_id = head_rel_name + '#' + chunk_idx。",
                    "例：Make_statement#2",
                    "用于在 JSONL 与 XLSX 间交叉定位某个 item。",
                )
            if field in {"head_rel_name", "chunk_idx"}:
                return (
                    "item 所属 head 与 chunk 序号。",
                    "chunk_idx 范围为 [0, E-1]。",
                    f"例：chunk_idx=0..{max(E-1,0)}",
                    "chunk_idx 可用于观察后段 chunk 是否更容易被丢弃（数据不足）。",
                )
            if field in {"num_paths", "num_rules"}:
                return (
                    "该 item 实际包含的 paths/rules 条数。",
                    "严格过滤开启时，应该恒等于 P/R。",
                    f"例：num_paths={P}, num_rules={R}",
                    "若不等，说明过滤逻辑或输入数据存在异常。",
                )
            if field in {"paths_conf_mean", "rules_conf_mean"}:
                return (
                    "该 item 内部的 conf 均值（paths 或 rules）。",
                    "对该 item chunk 内的 conf 做平均。",
                    "例：conf=[0.8,0.6] => mean=0.7",
                    "用于粗略衡量这个 item 的“质量”。",
                )
            if field.endswith("_body_atoms_mean"):
                return (
                    "该 item 内部的平均 body 原子数量（paths 或 rules）。",
                    "对该 item 中每条规则的 body_atoms 取均值。",
                    "例：两条规则 body_atoms=[2,3] => mean=2.5",
                    "值越大表示示例更复杂，也更占 prompt token。",
                )
            if field.endswith("_text_len_sum"):
                return (
                    "该 item 的文本长度总和（字符数）。",
                    "对 item.paths 或 item.rules 的字符串长度求和。",
                    "例：sum(len(rule_i))",
                    "用于估计 examples 部分占用的 prompt 长度（越大越占 token）。",
                )
            if field in {"path_tokens_total", "path_tokens_unique"}:
                return (
                    "该 item 的 token 数（总数/去重数），来自 item.paths 的 body。",
                    "path_tokens_total = 所有 path token 数；unique 为去重。",
                    "例：tokens=[A,B,A] => total=3, unique=2",
                    "unique 越大说明示例更“多样”；total 越大说明特征更“密”。",
                )

        if sheet == "Dropped":
            if field in {"head_rel_name", "chunk_idx"}:
                return (
                    "被丢弃 chunk 对的 head 与 chunk 序号。",
                    "chunk_idx 与 PerItem 的 chunk_idx 含义一致。",
                    "例：Make_statement, chunk_idx=4",
                    "用于定位具体丢弃发生在哪个 head 的第几个 chunk。",
                )
            if field in {"p_len", "r_len"}:
                return (
                    "被丢弃 chunk 的长度（paths/rules）。",
                    "p_len=len(p_chunk), r_len=len(r_chunk)。",
                    f"例：p_len=7(<P={P}) 导致 partial_paths",
                    "用来判断是“空”还是“差一点不满配”。",
                )
            if field == "reason":
                return (
                    "丢弃原因分类。",
                    "empty_* 表示某侧为 0；partial_* 表示某侧不足 P/R。",
                    "例：partial_paths / empty_rules",
                    "用于快速统计过滤损失来自哪一侧。",
                )

        if sheet == "TokenFreq":
            if field == "token":
                return (
                    "从 item.paths 的 rule body 抽取的关系名 token。",
                    "regex 抽取每个 R(...) 的 R（只取 body）。",
                    "例：A(X0,X1,T0) => token=A",
                    "这是检索相似度 TF-IDF 的基础特征。",
                )
            if field == "count_total":
                return (
                    "该 token 在所有保留 items 中出现的总次数。",
                    "对所有 item.paths 的 token 进行计数累加。",
                    example_int(300),
                    "值越大说明该 token 更常见，可能区分度更低。",
                )
            if field == "doc_freq_items":
                return (
                    "该 token 出现过的 item 数（文档频次）。",
                    "对每个 item 只计一次（set(token)），再对 item 数求和。",
                    example_int(120),
                    "用于理解 TF-IDF 的 DF：DF 越大 IDF 越小。",
                )

        # Fallback: ensure every field has at least a minimal explanation + example.
        lowered = field.lower()
        if "ratio" in lowered:
            ex = "例：0.05（表示 5%）"
        elif "conf" in lowered:
            ex = "例：0.75"
        elif "len" in lowered or "atoms" in lowered:
            ex = "例：3"
        elif lowered.startswith("n_") or "count" in lowered or "num" in lowered:
            ex = "例：10"
        else:
            ex = f"例：{field}=..."
        return ("字段说明（自动生成）", "见同名列/metric 的计算逻辑。", ex, "用于补全 Glossary，建议结合源码进一步确认。")

    header = ["sheet", "field", "meaning(解释)", "how_computed(计算方式)", "example(例子)", "how_to_read(解读)"]
    table: List[List[object]] = [header]

    def add(sheet: str, field: str) -> None:
        meaning, how, example, read = describe(sheet, field)
        table.append([sheet, field, meaning, how, example, read])

    # Summary metrics (one row per metric key)
    for metric in summary_metrics:
        add("Summary", metric)

    # Column dictionaries for the remaining sheets
    for col in per_head_columns:
        add("PerHead", col)
    for col in per_item_columns:
        add("PerItem", col)
    for col in dropped_columns:
        add("Dropped", col)
    for col in token_columns:
        add("TokenFreq", col)

    return table


def build_example_pool_items(
    sampled_paths_json: str,
    generated_rules_json: str,
    *,
    top_n_paths: int = 40,
    top_n_rules: int = 30,
    num_items_per_head: int = 5,
    paths_per_item: int = 6,
    rules_per_item: int = 6,
) -> List[dict]:
    """
    MVP pool construction:
    - For each head_rel_name:
      - take top-N sampled paths by conf
      - take top-N generated rules by conf
      - chunk them into E items with fixed sizes (P paths, R rules)
    """
    paths_by_head = _collect_rules_by_head(sampled_paths_json)
    rules_by_head = _collect_rules_by_head(generated_rules_json)

    common_heads = sorted(set(paths_by_head) & set(rules_by_head))
    pool_items: List[dict] = []
    for head_name in common_heads:
        top_paths = [r for _c, r in sorted(paths_by_head[head_name], key=lambda x: x[0], reverse=True)][
            :top_n_paths
        ]
        top_rules = [r for _c, r in sorted(rules_by_head[head_name], key=lambda x: x[0], reverse=True)][
            :top_n_rules
        ]

        path_chunks = _chunk_list(top_paths, paths_per_item, num_items_per_head)
        rule_chunks = _chunk_list(top_rules, rules_per_item, num_items_per_head)

        for idx, (p_chunk, r_chunk) in enumerate(zip(path_chunks, rule_chunks)):
            # Keep only items that have both inputs and outputs.
            if not p_chunk or not r_chunk:
                continue
            if len(p_chunk) < paths_per_item or len(r_chunk) < rules_per_item:
                continue
            tokens: List[str] = []
            for p in p_chunk:
                tokens.extend(extract_relation_tokens(p))
            pool_items.append(
                {
                    "ex_id": f"{head_name}#{idx}",
                    "head_rel_name": head_name,
                    "paths": p_chunk,
                    "rules": r_chunk,
                    "path_rels": tokens,
                }
            )
    return pool_items


def build_example_pool_items_and_stats(
    sampled_paths_json: str,
    generated_rules_json: str,
    *,
    top_n_paths: int = 40,
    top_n_rules: int = 30,
    num_items_per_head: int = 5,
    paths_per_item: int = 6,
    rules_per_item: int = 6,
) -> Tuple[List[dict], dict]:
    """
    Build JSONL items + a detailed stats dict for quality inspection.

    The stats focus on:
    - Coverage: how many heads/items are kept vs dropped
    - Completeness: strict (P,R) size requirements per item
    - Retrieval signal: token extraction richness from sampled paths
    """
    paths_by_head = _collect_rules_by_head(sampled_paths_json)
    rules_by_head = _collect_rules_by_head(generated_rules_json)

    heads_in_paths = set(paths_by_head)
    heads_in_rules = set(rules_by_head)
    common_heads = sorted(heads_in_paths & heads_in_rules)
    all_heads = sorted(heads_in_paths | heads_in_rules)

    required_paths = int(num_items_per_head) * int(paths_per_item)
    required_rules = int(num_items_per_head) * int(rules_per_item)

    per_head_rows: List[dict] = []
    per_item_rows: List[dict] = []
    dropped_rows: List[dict] = []

    token_total = Counter()
    token_df = Counter()

    pool_items: List[dict] = []

    for head_name in all_heads:
        in_paths = head_name in paths_by_head
        in_rules = head_name in rules_by_head

        paths_sorted = sorted(paths_by_head.get(head_name, []), key=lambda x: x[0], reverse=True)
        rules_sorted = sorted(rules_by_head.get(head_name, []), key=lambda x: x[0], reverse=True)

        n_paths_total = len(paths_sorted)
        n_rules_total = len(rules_sorted)
        n_paths_used = min(int(top_n_paths), n_paths_total)
        n_rules_used = min(int(top_n_rules), n_rules_total)

        paths_used = paths_sorted[:n_paths_used]
        rules_used = rules_sorted[:n_rules_used]

        # Used-level stats (after top-N truncation).
        paths_conf_used = [c for c, _r in paths_used]
        rules_conf_used = [c for c, _r in rules_used]
        paths_atoms_used = [_count_body_atoms(r) for _c, r in paths_used]
        rules_atoms_used = [_count_body_atoms(r) for _c, r in rules_used]

        used_tokens: List[str] = []
        used_empty_token_cnt = 0
        for _c, r in paths_used:
            toks = extract_relation_tokens(r)
            if not toks:
                used_empty_token_cnt += 1
            used_tokens.extend(toks)
        used_unique_tokens = len(set(used_tokens))
        used_empty_token_ratio = (used_empty_token_cnt / len(paths_used)) if paths_used else None

        items_non_empty = 0
        items_full = 0
        dropped_empty = 0
        dropped_partial = 0

        if head_name in common_heads:
            path_chunks = _chunk_list(paths_used, paths_per_item, num_items_per_head)
            rule_chunks = _chunk_list(rules_used, rules_per_item, num_items_per_head)

            for idx, (p_chunk_t, r_chunk_t) in enumerate(zip(path_chunks, rule_chunks)):
                p_len = len(p_chunk_t)
                r_len = len(r_chunk_t)
                if p_len > 0 and r_len > 0:
                    items_non_empty += 1

                # Keep only items that have both inputs and outputs.
                if p_len == 0 or r_len == 0:
                    dropped_empty += 1
                    reason = (
                        "empty_both"
                        if (p_len == 0 and r_len == 0)
                        else ("empty_paths" if p_len == 0 else "empty_rules")
                    )
                    dropped_rows.append(
                        {
                            "head_rel_name": head_name,
                            "chunk_idx": idx,
                            "p_len": p_len,
                            "r_len": r_len,
                            "reason": reason,
                        }
                    )
                    continue

                # Strict constraint: require full (P,R). (You added this for quality control.)
                if p_len < paths_per_item or r_len < rules_per_item:
                    dropped_partial += 1
                    if p_len < paths_per_item and r_len < rules_per_item:
                        reason = "partial_both"
                    elif p_len < paths_per_item:
                        reason = "partial_paths"
                    else:
                        reason = "partial_rules"
                    dropped_rows.append(
                        {
                            "head_rel_name": head_name,
                            "chunk_idx": idx,
                            "p_len": p_len,
                            "r_len": r_len,
                            "reason": reason,
                        }
                    )
                    continue

                items_full += 1
                p_rules = [r for _c, r in p_chunk_t]
                r_rules = [r for _c, r in r_chunk_t]

                tokens: List[str] = []
                for p in p_rules:
                    tokens.extend(extract_relation_tokens(p))

                token_total.update(tokens)
                token_df.update(set(tokens))

                pool_items.append(
                    {
                        "ex_id": f"{head_name}#{idx}",
                        "head_rel_name": head_name,
                        "paths": p_rules,
                        "rules": r_rules,
                        "path_rels": tokens,
                    }
                )

                per_item_rows.append(
                    {
                        "ex_id": f"{head_name}#{idx}",
                        "head_rel_name": head_name,
                        "chunk_idx": idx,
                        "num_paths": len(p_rules),
                        "num_rules": len(r_rules),
                        "paths_conf_mean": _safe_mean([c for c, _r in p_chunk_t]),
                        "rules_conf_mean": _safe_mean([c for c, _r in r_chunk_t]),
                        "paths_body_atoms_mean": _safe_mean([float(_count_body_atoms(r)) for _c, r in p_chunk_t]),
                        "rules_body_atoms_mean": _safe_mean([float(_count_body_atoms(r)) for _c, r in r_chunk_t]),
                        "paths_text_len_sum": sum(len(r) for r in p_rules),
                        "rules_text_len_sum": sum(len(r) for r in r_rules),
                        "path_tokens_total": len(tokens),
                        "path_tokens_unique": len(set(tokens)),
                    }
                )

        # Determine bottleneck per head to explain missing items.
        if not in_paths:
            bottleneck = "missing_paths"
        elif not in_rules:
            bottleneck = "missing_rules"
        else:
            full_path_chunks = (n_paths_used // int(paths_per_item)) if paths_per_item else 0
            full_rule_chunks = (n_rules_used // int(rules_per_item)) if rules_per_item else 0
            if min(full_path_chunks, full_rule_chunks) >= int(num_items_per_head):
                bottleneck = "ok_full"
            elif full_path_chunks < full_rule_chunks:
                bottleneck = "paths_short"
            elif full_rule_chunks < full_path_chunks:
                bottleneck = "rules_short"
            else:
                bottleneck = "both_short"

        per_head_rows.append(
            {
                "head_rel_name": head_name,
                "in_paths": in_paths,
                "in_rules": in_rules,
                "n_paths_total": n_paths_total,
                "n_rules_total": n_rules_total,
                "n_paths_used": n_paths_used,
                "n_rules_used": n_rules_used,
                "required_paths(E*P)": required_paths,
                "required_rules(E*R)": required_rules,
                "full_path_chunks": (n_paths_used // int(paths_per_item)) if paths_per_item else 0,
                "full_rule_chunks": (n_rules_used // int(rules_per_item)) if rules_per_item else 0,
                "items_non_empty": items_non_empty if head_name in common_heads else 0,
                "items_full_written": items_full if head_name in common_heads else 0,
                "items_target_E": int(num_items_per_head) if head_name in common_heads else 0,
                "dropped_empty": dropped_empty if head_name in common_heads else 0,
                "dropped_partial": dropped_partial if head_name in common_heads else 0,
                "bottleneck": bottleneck,
                "paths_conf_mean_used": _safe_mean(paths_conf_used),
                "paths_conf_median_used": _safe_median(paths_conf_used),
                "paths_conf_min_used": _safe_min(paths_conf_used),
                "paths_conf_max_used": _safe_max(paths_conf_used),
                "rules_conf_mean_used": _safe_mean(rules_conf_used),
                "rules_conf_median_used": _safe_median(rules_conf_used),
                "rules_conf_min_used": _safe_min(rules_conf_used),
                "rules_conf_max_used": _safe_max(rules_conf_used),
                "paths_body_atoms_mean_used": _safe_mean([float(x) for x in paths_atoms_used]) if paths_atoms_used else None,
                "rules_body_atoms_mean_used": _safe_mean([float(x) for x in rules_atoms_used]) if rules_atoms_used else None,
                "paths_unique_tokens_used": used_unique_tokens,
                "paths_empty_token_ratio_used": used_empty_token_ratio,
            }
        )

    # Summary table (2 columns: metric/value)
    head_items = [r["items_full_written"] for r in per_head_rows if r["head_rel_name"] in common_heads]
    hist = Counter(head_items)
    heads_with_ge1 = sum(1 for x in head_items if x >= 1)
    heads_with_fullE = sum(1 for x in head_items if x >= int(num_items_per_head))
    heads_paths_used_ge_req = sum(1 for r in per_head_rows if r["head_rel_name"] in common_heads and r["n_paths_used"] >= required_paths)
    heads_rules_used_ge_req = sum(1 for r in per_head_rows if r["head_rel_name"] in common_heads and r["n_rules_used"] >= required_rules)
    heads_both_used_ge_req = sum(
        1
        for r in per_head_rows
        if r["head_rel_name"] in common_heads and r["n_paths_used"] >= required_paths and r["n_rules_used"] >= required_rules
    )

    summary_pairs = [
        ("sampled_paths_json", sampled_paths_json),
        ("generated_rules_json", generated_rules_json),
        ("top_n_paths", top_n_paths),
        ("top_n_rules", top_n_rules),
        ("num_items_per_head(E)", num_items_per_head),
        ("paths_per_item(P)", paths_per_item),
        ("rules_per_item(R)", rules_per_item),
        ("required_paths(E*P)", required_paths),
        ("required_rules(E*R)", required_rules),
        ("heads_in_paths", len(heads_in_paths)),
        ("heads_in_rules", len(heads_in_rules)),
        ("heads_intersection", len(common_heads)),
        ("candidate_items_total(heads*E)", len(common_heads) * int(num_items_per_head)),
        ("items_written", len(pool_items)),
        ("items_dropped_empty", sum(r.get("dropped_empty", 0) for r in per_head_rows)),
        ("items_dropped_partial", sum(r.get("dropped_partial", 0) for r in per_head_rows)),
        ("heads_with>=1_item", heads_with_ge1),
        ("heads_with_full_E_items", heads_with_fullE),
        ("heads_paths_used>=E*P", heads_paths_used_ge_req),
        ("heads_rules_used>=E*R", heads_rules_used_ge_req),
        ("heads_both_used>=E*P_and_E*R", heads_both_used_ge_req),
        ("items_per_head_mean", _safe_mean([float(x) for x in head_items]) if head_items else None),
        ("items_per_head_median", _safe_median([float(x) for x in head_items]) if head_items else None),
        ("unique_tokens_in_kept_items", len(token_total)),
    ]
    for i in range(0, int(num_items_per_head) + 1):
        summary_pairs.append((f"heads_with_items={i}", hist.get(i, 0)))
    summary_table: List[List[object]] = [["metric", "value"]] + [[k, v] for k, v in summary_pairs]

    token_rows = [
        {"token": tok, "count_total": cnt, "doc_freq_items": token_df.get(tok, 0)}
        for tok, cnt in token_total.most_common(2000)
    ]

    stats = {
        "summary_table": summary_table,
        "per_head_rows": per_head_rows,
        "per_item_rows": per_item_rows,
        "dropped_rows": dropped_rows,
        "token_rows": token_rows,
        "common_heads": common_heads,
    }
    return pool_items, stats


def write_jsonl(items: Iterable[dict], out_path: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        for item in items:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build example_pool.jsonl for dynamic few-shot retrieval.")
    parser.add_argument("--dataset", type=str, required=True, help="Dataset name, e.g., icews14")
    parser.add_argument(
        "--sampled_paths_json",
        type=str,
        default=None,
        help="Path to sampled confidence_concrete.json (rule_sampler output).",
    )
    parser.add_argument(
        "--generated_rules_json",
        type=str,
        default=None,
        help="Path to 0_confidence_concrete.json (Iteration_reasoning first iteration output).",
    )
    parser.add_argument(
        "--out",
        type=str,
        default=None,
        help="Output JSONL path. A companion <out_basename>_stats.xlsx will be written next to it.",
    )
    parser.add_argument("--top_n_paths", type=int, default=40)
    parser.add_argument("--top_n_rules", type=int, default=30)
    parser.add_argument("--num_items_per_head", type=int, default=5, help="E: number of items per head relation")
    parser.add_argument("--paths_per_item", type=int, default=6, help="P: number of sampled paths per item")
    parser.add_argument("--rules_per_item", type=int, default=6, help="R: number of generated rules per item")
    parser.add_argument("--bat_file_name", type=str, default='bat_file',
                        help="Batch file name")
    parser.add_argument("--results_root_path", type=str, default='results',
                        help="Results root path. Must put this parameter at last position on cmd line to avoid parsing error.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    args.results_root_path = args.results_root_path.strip('"')
    
    sampled_paths_json = args.sampled_paths_json or maybe_windows_long_path(os.path.join( args.results_root_path, args.bat_file_name,
        "sampled_path", args.dataset, "confidence_concrete.json"
    ))

    generated_rules_json = args.generated_rules_json or maybe_windows_long_path(os.path.join( args.results_root_path, args.bat_file_name,
        "gen_rules_iteration_for_example_pool", args.dataset, "evaluation", "train", "0_confidence_concrete.json"
    ))


    out_dir_to_remove =  maybe_windows_long_path(os.path.join(args.results_root_path, args.bat_file_name, "example_pool", args.dataset))
    
    if os.path.exists(out_dir_to_remove):
        shutil.rmtree(out_dir_to_remove, ignore_errors=False)
    os.makedirs(out_dir_to_remove, exist_ok=True)   
    
    args.out = args.out or maybe_windows_long_path(os.path.join(args.results_root_path, args.bat_file_name, "example_pool", args.dataset, "example_pool.jsonl"))

    if not os.path.exists(sampled_paths_json):
        raise FileNotFoundError(f"sampled_paths_json not found: {sampled_paths_json}")
    if not os.path.exists(generated_rules_json):
        raise FileNotFoundError(f"generated_rules_json not found: {generated_rules_json}")

    items, stats = build_example_pool_items_and_stats(
        sampled_paths_json,
        generated_rules_json,
        top_n_paths=args.top_n_paths,
        top_n_rules=args.top_n_rules,
        num_items_per_head=args.num_items_per_head,
        paths_per_item=args.paths_per_item,
        rules_per_item=args.rules_per_item,
    )
    write_jsonl(items, args.out)
    print(f"Wrote {len(items)} example items to: {args.out}")

    # Write detailed build statistics as an XLSX next to args.out.
    out_dir = os.path.dirname(os.path.abspath(args.out)) or "."
    out_base = os.path.splitext(os.path.basename(args.out))[0]
    stats_path = maybe_windows_long_path(os.path.join(out_dir, f"{out_base}_stats.xlsx"))

    # Enrich the summary with run metadata.
    summary_table = stats["summary_table"]
    summary_table.insert(1, ["dataset", args.dataset])
    summary_table.insert(2, ["out_jsonl", args.out])
    summary_table.insert(3, ["stats_xlsx", stats_path])
    summary_table.insert(4, ["strict_full_chunk_filter", True])
    summary_table.insert(5, ["token_extraction_body_only", True])

    per_head_columns = [
        "head_rel_name",
        "in_paths",
        "in_rules",
        "n_paths_total",
        "n_rules_total",
        "n_paths_used",
        "n_rules_used",
        "required_paths(E*P)",
        "required_rules(E*R)",
        "full_path_chunks",
        "full_rule_chunks",
        "items_non_empty",
        "items_full_written",
        "items_target_E",
        "dropped_empty",
        "dropped_partial",
        "bottleneck",
        "paths_conf_mean_used",
        "paths_conf_median_used",
        "paths_conf_min_used",
        "paths_conf_max_used",
        "rules_conf_mean_used",
        "rules_conf_median_used",
        "rules_conf_min_used",
        "rules_conf_max_used",
        "paths_body_atoms_mean_used",
        "rules_body_atoms_mean_used",
        "paths_unique_tokens_used",
        "paths_empty_token_ratio_used",
    ]
    per_item_columns = [
        "ex_id",
        "head_rel_name",
        "chunk_idx",
        "num_paths",
        "num_rules",
        "paths_conf_mean",
        "rules_conf_mean",
        "paths_body_atoms_mean",
        "rules_body_atoms_mean",
        "paths_text_len_sum",
        "rules_text_len_sum",
        "path_tokens_total",
        "path_tokens_unique",
    ]
    dropped_columns = ["head_rel_name", "chunk_idx", "p_len", "r_len", "reason"]
    token_columns = ["token", "count_total", "doc_freq_items"]

    summary_metrics = [row[0] for row in summary_table[1:]]
    glossary_table = build_glossary_table(
        summary_metrics=summary_metrics,
        per_head_columns=per_head_columns,
        per_item_columns=per_item_columns,
        dropped_columns=dropped_columns,
        token_columns=token_columns,
        E=int(args.num_items_per_head),
        P=int(args.paths_per_item),
        R=int(args.rules_per_item),
    )

    write_stats_xlsx(
        [
            ("Summary", summary_table),
            ("PerHead", _to_table(stats["per_head_rows"], per_head_columns)),
            ("PerItem", _to_table(stats["per_item_rows"], per_item_columns)),
            ("Dropped", _to_table(stats["dropped_rows"], dropped_columns)),
            ("TokenFreq", _to_table(stats["token_rows"], token_columns)),
            ("Glossary", glossary_table),
        ],
        stats_path,
    )
    print(f"Wrote stats XLSX to: {stats_path}")


if __name__ == "__main__":
    main()
