# evaluate.py
import csv
import json
import os
import sys
import numpy as np

from grapher import Grapher
from params import get_params
from utils_windows_long_path import maybe_windows_long_path
from vlrg_utils import candidates_run_id, fmt_float_for_name

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


# -----------------------------
# Original helpers (kept logic)
# -----------------------------
def load_candidates(ranked_rules_dir, candidates_file):
    with open(maybe_windows_long_path(os.path.join(ranked_rules_dir, candidates_file)), "r", encoding="utf-8") as f:
        candidates = json.load(f)
    # candidates: {i: {entity: score}}
    return {int(k): {int(cand): v for cand, v in vv.items()} for k, vv in candidates.items()}


def build_test_index(test_numpy):
    index = {}
    for i, row in enumerate(test_numpy):
        key = tuple(int(x) for x in row)
        if key not in index:
            index[key] = i
    return index


def load_test_and_score_data(dataset, dataset_dir, graph_reasoning_type, split):
    if split == "valid":
        test_path = maybe_windows_long_path(os.path.join(dataset_dir, graph_reasoning_type, "test_valid.npy"))
        score_path = maybe_windows_long_path(os.path.join(dataset_dir, graph_reasoning_type, "score_valid.npy"))
        if not os.path.exists(score_path):
            typo_score_path = maybe_windows_long_path(os.path.join(dataset_dir, graph_reasoning_type, "score_vlid.npy"))
            if os.path.exists(typo_score_path):
                score_path = typo_score_path
    else:
        test_path = maybe_windows_long_path(os.path.join(dataset_dir, graph_reasoning_type, "test.npy"))
        score_path = maybe_windows_long_path(os.path.join(dataset_dir, graph_reasoning_type, "score.npy"))

    missing = [p for p in (test_path, score_path) if not os.path.exists(p)]
    if missing:
        raise FileNotFoundError(
            "Missing graph reasoning baseline files required by evaluate.py:\n"
            f"  {test_path}\n"
            f"  {score_path}\n"
            "Export tips:\n"
            "  - TiRGN: run TiRGN/src/main.py with `--test --export-npy --export-npy-dir <this_folder>`\n"
            "          (add `--test-split valid` when evaluating valid)\n"
            "  - LogGL (LogCL): run LogCL/src/main.py training with `--export-npy --export-npy-dir <this_folder>`\n"
            "                (it exports both test+valid baselines after training)"
        )

    test_numpy = np.load(test_path)
    if dataset == "icews18":
        test_numpy[:, 3] = (test_numpy[:, 3] / 24).astype(int)
    score_numpy = np.load(score_path)
    return test_numpy, score_numpy


# -----------------------------
# Fast multi-weight evaluation
# -----------------------------
def build_other_answers_map(test_data: np.ndarray):
    """
    Match your filter_candidates():
    For each key (sub, rel, ts), collect all objects.
    In evaluation for a query (s,r,o,t), we will filter out objects != o under same key.
    """
    mp = {}
    # test_data columns assumed: [sub, rel, obj, ts]
    for row in test_data:
        s = int(row[0])
        r = int(row[1])
        o = int(row[2])
        t = int(row[3])
        key = (s, r, t)
        if key not in mp:
            mp[key] = []
        mp[key].append(o)
    return mp


def get_scores_arrays_for_query(
    test_query: np.ndarray,
    i: int,
    num_entities: int,
    all_rule_candidates: dict,
    test_numpy: np.ndarray,
    score_numpy: np.ndarray,
    test_index: dict,
):
    """
    Return regcn_score (N,), rule_score (N,) float32
    """
    # rule vector
    rule_score = np.zeros(num_entities, dtype=np.float32)
    rc = all_rule_candidates.get(int(i), {})  # dict: entity -> score
    if rc:
        idx = np.fromiter(rc.keys(), dtype=np.int64)
        val = np.fromiter(rc.values(), dtype=np.float32)
        rule_score[idx] = val

    # regcn vector
    if test_index is not None:
        pos = test_index.get(tuple(int(x) for x in test_query))
        if pos is None:
            raise KeyError(f"Query not found in baseline test.npy: {test_query}")
        regcn_score = score_numpy[pos].astype(np.float32, copy=False)
    else:
        indices = np.where((test_numpy == test_query).all(axis=1))[0]
        if len(indices) == 0:
            raise KeyError(f"Query not found in baseline test.npy (slow path): {test_query}")
        regcn_score = score_numpy[indices[0]].astype(np.float32, copy=False)

    return regcn_score, rule_score


def calculate_ranks_multi(scores_w_n: np.ndarray, ans: int, setting: str = "best") -> np.ndarray:
    """
    Equivalent to your calculate_rank() but vectorized over multiple weights.
    scores_w_n: shape (W, N), already with filtered entities set to -inf.
    ans: int
    return ranks: (W,)
    """
    # conf for answer under each weight
    conf = scores_w_n[:, ans]  # (W,)

    # strictly greater counts => best rank
    greater = (scores_w_n > conf[:, None]).sum(axis=1)  # (W,)

    # exact equality counts (to match your original x == conf behavior)
    equal = (scores_w_n == conf[:, None]).sum(axis=1)   # (W,)

    if setting == "best":
        return greater + 1
    elif setting == "worst":
        return greater + equal
    elif setting == "average":
        best = greater + 1
        worst = greater + equal
        return (best + worst) // 2
    else:
        raise ValueError(f"Unknown setting: {setting}")


def save_results_to_excel(out_dir, candidates_file_name, graph_reasoning_type, rows, parsed, mr_09, file_suffix=""):
    if Workbook is None:
        print("[Warn] openpyxl not installed, skip excel output.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "eval"

    header = ["candidates_file", "graph_reasoning_type", "weight", "Hits@1", "Hits@3", "Hits@10", "MRR", "num_queries"]
    ws.append(header)

    for r in rows:
        ws.append([
            candidates_file_name,
            graph_reasoning_type,
            r["weight"],
            r["hits1"],
            r["hits3"],
            r["hits10"],
            r["mrr"],
            r["num_queries"],
        ])

    for col in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    run_id = candidates_run_id(candidates_file_name)
    xlsx_name = f"weights_eval_{run_id}_{graph_reasoning_type}_{file_suffix}.xlsx"
    xlsx_path = maybe_windows_long_path(os.path.join(out_dir, xlsx_name))
    wb.save(xlsx_path)
    print(f"[Saved] Excel => {xlsx_path}")


def _safe_write_json(path: str, obj: dict):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(maybe_windows_long_path(path), "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


def _safe_write_csv(path: str, rows: list):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(maybe_windows_long_path(path), "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["weight", "hits1", "hits3", "hits10", "mrr", "num_queries"])
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


def main():
    parsed = get_params()

    dataset = parsed.dataset
    candidates_file_name = parsed.candidates_file_name

    dataset_dir = maybe_windows_long_path(os.path.join("datasets", dataset))
    parsed.results_root_path = parsed.results_root_path.strip('"')
    ranked_rules_dir = maybe_windows_long_path(
        os.path.join(parsed.results_root_path, parsed.bat_file_name, "ranked_rules", dataset)
    )

    data = Grapher(dataset_dir)
    num_entities = len(data.id2entity)
    test_data = data.test_idx if (parsed.test_data == "test") else data.valid_idx

    # Optional: evaluate only forward (non-inverse) queries.
    query_dir = str(getattr(parsed, "query_dir", "both") or "both").lower().strip()
    if query_dir not in ("both", "forward"):
        raise ValueError(f"Invalid --query_dir: {query_dir} (expected both|forward)")

    # Added for subset prompt-eval (2026-01-20)
    # Optional: compute extra metrics on a selected subset of (folded) relations.
    num_relations_old = int(len(getattr(data, "relation2id_old", {})))
    keep_ids = None
    selected_relations_path = maybe_windows_long_path(
        os.path.join(parsed.results_root_path, parsed.bat_file_name, "sampled_path", dataset, "selected_relations.json")
    )
    if parsed.selected_relations is True and os.path.exists(selected_relations_path):
        with open(selected_relations_path, "r", encoding="utf-8") as f:
            selected_payload = json.load(f) or {}
        keep_ids = set(int(x) for x in (selected_payload.get("selected_head_rel_ids", []) or []))
        print(f"[Subset] Loaded keep_ids={len(keep_ids)} from {selected_relations_path}")

    all_rule_candidates = load_candidates(ranked_rules_dir, candidates_file_name)

    # 0.00, 0.05, ..., 1.00  （共 21 个）
    weights = np.arange(0.0, 1.0 + 1e-9, 0.05, dtype=np.float32)
    # 防止浮点显示成 0.15000001 之类，输出/文件名更干净
    weights = np.round(weights, 2)


    # If not TiRGN/REGCN/LogGL: rule-only; weight doesn't change anything (still output for completeness)
    use_graph = parsed.graph_reasoning_type in ["TiRGN", "REGCN", "LogGL"]

    if use_graph:
        test_numpy, score_numpy = load_test_and_score_data(dataset, dataset_dir, parsed.graph_reasoning_type, parsed.test_data)
        test_index = build_test_index(test_numpy)
        other_answers_map = build_other_answers_map(test_data)
    else:
        test_numpy = score_numpy = test_index = None
        other_answers_map = None

    # metrics accumulators per weight
    W = len(weights)
    hits1 = np.zeros(W, dtype=np.int64)
    hits3 = np.zeros(W, dtype=np.int64)
    hits10 = np.zeros(W, dtype=np.int64)
    mrr = np.zeros(W, dtype=np.float64)

    # Added for subset prompt-eval (2026-01-20)
    subset_hits1 = np.zeros(W, dtype=np.int64)
    subset_hits3 = np.zeros(W, dtype=np.int64)
    subset_hits10 = np.zeros(W, dtype=np.int64)
    subset_mrr = np.zeros(W, dtype=np.float64)
    subset_num_samples = 0

    setting = getattr(parsed, "rank_setting", "best")  # optionally allow parsed.rank_setting

    if query_dir == "forward":
        if num_relations_old <= 0:
            print("[Warn] query_dir=forward but relation2id_old missing; fall back to both.")
            eval_qids = list(range(len(test_data)))
        else:
            eval_qids = np.nonzero(test_data[:, 1] < num_relations_old)[0].astype(int).tolist()
        print(f"[QueryDir] query_dir=forward => eval_qids={len(eval_qids)}/{len(test_data)}")
    else:
        eval_qids = list(range(len(test_data)))

    num_samples = len(eval_qids)

    iterator = eval_qids
    if tqdm is not None:
        iterator = tqdm(iterator, total=num_samples, desc="Evaluating", file=sys.stdout)

    for i in iterator:
        test_query = test_data[i]
        s = int(test_query[0])
        r = int(test_query[1])
        ans = int(test_query[2])
        t = int(test_query[3])

        # Added for subset prompt-eval (2026-01-20)
        is_subset = False
        if keep_ids is not None and num_relations_old > 0:
            if (int(r) % int(num_relations_old)) in keep_ids:
                is_subset = True

        if use_graph:
            regcn_score, rule_score = get_scores_arrays_for_query(
                test_query=test_query,
                i=i,
                num_entities=num_entities,
                all_rule_candidates=all_rule_candidates,
                test_numpy=test_numpy,
                score_numpy=score_numpy,
                test_index=test_index,
            )

            # combined: (W, N) = regcn + w*(rule - regcn)
            delta = rule_score - regcn_score
            scores = regcn_score[None, :] + weights[:, None] * delta[None, :]

            # filter (equivalent to your filter_candidates pop)
            key = (s, r, t)
            objs = other_answers_map.get(key, [])
            if objs:
                # set filtered objects (obj != ans) to -inf for all weights
                # but keep ans
                for o in objs:
                    if int(o) != ans:
                        scores[:, int(o)] = -np.inf

            # ranks for all weights at once
            ranks = calculate_ranks_multi(scores, ans, setting=setting).astype(np.int64)

        else:
            # rule-only: candidates dict; run your original logic (weight irrelevant)
            # We'll compute rank once and broadcast.
            candidates = (all_rule_candidates.get(int(i), {}) or {}).copy()

            # apply the same filter_candidates logic
            other_answers = test_data[
                (test_data[:, 0] == test_query[0])
                * (test_data[:, 1] == test_query[1])
                * (test_data[:, 2] != test_query[2])
                * (test_data[:, 3] == test_query[3])
            ]
            if len(other_answers):
                objects = other_answers[:, 2]
                for obj in objects:
                    candidates.pop(int(obj), None)

            # calculate_rank(setting="best") equivalent
            rank = num_entities
            if ans in candidates:
                conf = candidates[ans]
                all_confs = sorted(list(candidates.values()), reverse=True)
                ranks_same = [idx for idx, x in enumerate(all_confs) if x == conf]
                if setting == "average":
                    rank = (ranks_same[0] + ranks_same[-1]) // 2 + 1
                elif setting == "best":
                    rank = ranks_same[0] + 1
                elif setting == "worst":
                    rank = ranks_same[-1] + 1
                else:
                    raise ValueError(f"Unknown setting: {setting}")

            ranks = np.full(W, int(rank), dtype=np.int64)

        # update metrics vectorized
        hits10 += (ranks <= 10).astype(np.int64)
        hits3 += (ranks <= 3).astype(np.int64)
        hits1 += (ranks == 1).astype(np.int64)
        mrr += 1.0 / ranks

        # Added for subset prompt-eval (2026-01-20)
        if is_subset:
            subset_num_samples += 1
            subset_hits10 += (ranks <= 10).astype(np.int64)
            subset_hits3 += (ranks <= 3).astype(np.int64)
            subset_hits1 += (ranks == 1).astype(np.int64)
            subset_mrr += 1.0 / ranks

    # normalize
    hits1_f = hits1 / num_samples
    hits3_f = hits3 / num_samples
    hits10_f = hits10 / num_samples
    mrr_f = mrr / num_samples

    # Added for subset prompt-eval (2026-01-20)
    coverage = float(subset_num_samples) / float(num_samples) if num_samples > 0 else 0.0
    if subset_num_samples > 0:
        subset_hits1_f = subset_hits1 / subset_num_samples
        subset_hits3_f = subset_hits3 / subset_num_samples
        subset_hits10_f = subset_hits10 / subset_num_samples
        subset_mrr_f = subset_mrr / subset_num_samples
    else:
        subset_hits1_f = np.zeros(W, dtype=np.float64)
        subset_hits3_f = np.zeros(W, dtype=np.float64)
        subset_hits10_f = np.zeros(W, dtype=np.float64)
        subset_mrr_f = np.zeros(W, dtype=np.float64)

    mr_09 = 0.0
    suset_mr_09 = 0.0
    # print + save
    excel_rows = []
    subset_excel_rows = []
    for wi, w in enumerate(weights.tolist()):
        h1 = float(hits1_f[wi])
        h3 = float(hits3_f[wi])
        h10 = float(hits10_f[wi])
        mr = float(mrr_f[wi])
        if wi == 18:
            mr_09 = round(mr, 6)
        print(f"\n===== graph={parsed.graph_reasoning_type} | rule_weight={w} | setting={setting} =====")
        print("Hits@1: ", round(h1, 6))
        print("Hits@3: ", round(h3, 6))
        print("Hits@10:", round(h10, 6))
        print("MRR:    ", round(mr, 6))

        # Added for subset prompt-eval (2026-01-20)
        if keep_ids is not None:
            sh1 = float(subset_hits1_f[wi])
            sh3 = float(subset_hits3_f[wi])
            sh10 = float(subset_hits10_f[wi])
            smr = float(subset_mrr_f[wi])
            if wi == 18:
                suset_mr_09 = round(smr, 6)
            print(f"[SUBSET] num_samples={subset_num_samples} coverage={coverage:.6f}")
            print("[SUBSET] Hits@1: ", round(sh1, 6))
            print("[SUBSET] Hits@3: ", round(sh3, 6))
            print("[SUBSET] Hits@10:", round(sh10, 6))
            print("[SUBSET] MRR:    ", round(smr, 6))


        excel_rows.append({
            "weight": float(w),
            "hits1": round(h1, 6),
            "hits3": round(h3, 6),
            "hits10": round(h10, 6),
            "mrr": round(mr, 6),
            "num_queries": int(num_samples),
        })
        if keep_ids is not None:
            subset_excel_rows.append({
                "weight": float(w),
                "hits1": round(sh1, 6),
                "hits3": round(sh3, 6),
                "hits10": round(sh10, 6),
                "mrr": round(smr, 6),
                "num_queries": int(subset_num_samples),
            })
        

    save_results_to_excel(
        out_dir=ranked_rules_dir,
        candidates_file_name=candidates_file_name,
        graph_reasoning_type=parsed.graph_reasoning_type,
        rows=excel_rows,
        parsed=parsed,
        mr_09=mr_09,
        file_suffix = "full",
    )
    if keep_ids is not None:
        save_results_to_excel(
            out_dir=ranked_rules_dir,
            candidates_file_name=candidates_file_name,
            graph_reasoning_type=parsed.graph_reasoning_type,
            rows=subset_excel_rows,
            parsed=parsed,
            mr_09=suset_mr_09,
            file_suffix = "subset",
        )

    # -----------------------------
    # VLRG: always save JSON/CSV metrics for the requested rule_weight (default 0.9)
    # -----------------------------
    cand_base = os.path.splitext(os.path.basename(candidates_file_name))[0]
    w_req = float(getattr(parsed, "rule_weight", 0.9))
    w_req = float(np.round(w_req, 2))
    # Find closest weight in grid
    weights_list = [float(x["weight"]) for x in excel_rows]
    best_i = min(range(len(weights_list)), key=lambda i: abs(weights_list[i] - w_req)) if weights_list else 0
    picked = excel_rows[best_i] if excel_rows else {}

    metrics_payload = {
        "meta": {
            "dataset": dataset,
            "split": parsed.test_data,
            "graph_reasoning_type": parsed.graph_reasoning_type,
            "rule_weight": w_req,
            "query_dir": query_dir,
            "candidates_file": candidates_file_name,
        },
        "metrics": picked,
        "grid": excel_rows,
    }
    run_id = candidates_run_id(candidates_file_name)
    w_str = fmt_float_for_name(w_req, decimals=2)
    out_json = maybe_windows_long_path(os.path.join(ranked_rules_dir, f"metrics_{run_id}_{parsed.graph_reasoning_type}_w{w_str}.json"))
    out_csv = maybe_windows_long_path(os.path.join(ranked_rules_dir, f"metrics_{run_id}_{parsed.graph_reasoning_type}_w{w_str}.csv"))
    _safe_write_json(out_json, metrics_payload)
    _safe_write_csv(out_csv, [picked] if picked else [])
    print(f"[Saved] metrics json => {maybe_windows_long_path(out_json)}")
    print(f"[Saved] metrics csv  => {maybe_windows_long_path(out_csv)}")

if __name__ == "__main__":
    main()
